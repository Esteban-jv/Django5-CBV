from django.shortcuts import render
import xlsxwriter
import openpyxl

# Create your views here.
def excel_writer(request):
    filename = "documents/LibroExcel3.xlsx"

    data = [
        ["Name", "Surname", "Age"],
        ["Jon", "Snow", 33],
        ["Daenerys", "Targaryen", 25],
        ["Tyrion", "Lannister", 40],
        ["Jaime", "Lannister", 35],
        ["Cersei", "Lannister", 36]
    ]

    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet("Hojax 1")

    for row in range(len(data)):
        for col in range(len(data[row])):
            worksheet.write(row, col, data[row][col])

    workbook.close()

    return render(request, "csv.html")

def excel_reader(request):
    filename = "documents/LibroExcel.xlsx"

    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    for row in range(worksheet.max_row):
        for col in worksheet.iter_cols(row, worksheet.max_column):
            print( col[row].value, end="\t")
            print('')

    workbook.close()
    return render(request, "csv.html")